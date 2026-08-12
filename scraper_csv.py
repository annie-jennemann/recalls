"""Download the FDA recalls workbook and save it as a CSV file."""

import csv
import html
import random
import tempfile
from datetime import date, datetime
from pathlib import Path
from urllib.parse import urljoin

from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright


FDA_URL = (
    "https://www.fda.gov/"
    "safety/recalls-market-withdrawals-safety-alerts"
)

OUTPUT_PATH = Path("data/fda_recalls.csv")


def download_workbook(destination: Path) -> None:
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(
            accept_downloads=True,
            user_agent=(
                "Mozilla/5.0 (X11; Linux x86_64) "
                "AppleWebKit/537.36 "
                "(KHTML, like Gecko) "
                "Chrome/131 Safari/537.36"
            ),
        )

        try:
            page = context.new_page()
            page.goto(
                FDA_URL,
                wait_until="domcontentloaded",
                timeout=60_000,
            )

            link = page.locator(
                'a[href*="datatables-data"][href*="_format=xlsx"]'
            ).first

            try:
                href = link.get_attribute(
                    "href",
                    timeout=30_000,
                )
            except PlaywrightTimeoutError:
                href = None

            if href:
                download_url = urljoin(
                    FDA_URL,
                    html.unescape(href),
                )

                with page.expect_download(
                    timeout=60_000
                ) as download_info:
                    link.click(force=True)

                download_info.value.save_as(
                    str(destination)
                )

                print(f"Downloaded: {download_url}")

            else:
                download_url = (
                    f"{FDA_URL}/datatables-data"
                    f"?randparam={random.randint(100000, 999999)}"
                    "&page&_format=xlsx"
                )

                with page.expect_download(
                    timeout=60_000
                ) as download_info:
                    page.goto(
                        download_url,
                        wait_until="commit",
                        timeout=60_000,
                    )

                download_info.value.save_as(
                    str(destination)
                )

                print(f"Downloaded fallback URL: {download_url}")

        finally:
            context.close()
            browser.close()


def normalize_cell(value: object) -> object:
    if value is None:
        return ""

    if isinstance(value, (datetime, date)):
        return value.isoformat()

    return value


def workbook_to_csv(
    workbook_path: Path,
    csv_path: Path,
) -> None:
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    try:
        worksheet = workbook.active
        csv_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        with csv_path.open(
            "w",
            newline="",
            encoding="utf-8",
        ) as output:
            writer = csv.writer(output)

            for row in worksheet.iter_rows(
                values_only=True
            ):
                writer.writerow(
                    [
                        normalize_cell(value)
                        for value in row
                    ]
                )

    finally:
        workbook.close()


def main() -> None:
    OUTPUT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with tempfile.TemporaryDirectory() as temp_dir:
        workbook_path = (
            Path(temp_dir) / "fda-recalls.xlsx"
        )
        csv_path = (
            Path(temp_dir) / "fda_recalls.csv"
        )

        download_workbook(workbook_path)
        workbook_to_csv(workbook_path, csv_path)

        new_contents = csv_path.read_bytes()

        old_contents = (
            OUTPUT_PATH.read_bytes()
            if OUTPUT_PATH.exists()
            else None
        )

        OUTPUT_PATH.write_bytes(new_contents)

    if new_contents == old_contents:
        print("CSV unchanged.")
    else:
        print(f"CSV updated: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
